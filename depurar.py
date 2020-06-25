#DESARROLLADO POR JUAN CARLOS AGUIRRE 2019

from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.utils.cell import get_column_letter

import csv
import sys
import time

import pandas as pd

read_file = pd.read_csv (r'./ACTIVIDAD.csv')
read_file.to_excel (r'./ACTIVIDAD.xlsx', index = None, header=True)

#dest_filename = 'ACTIVIDAD.xlsx'
#wb = Workbook()
#ws = wb.active
#ws.title = 'PISIS'
#with open('ACTIVIDAD.csv', "r") as f:
#    reader = csv.reader(f, delimiter=',')
#    for row in reader:
#        ws.append(row)
#        print (row)

#wb.save(filename = dest_filename)

